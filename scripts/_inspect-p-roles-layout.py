"""Inspect P-Roles worksheet + PivotTable OOXML without saving the workbook."""
from __future__ import annotations

import hashlib
import json
import sys
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def inspect(path: str) -> dict:
    out: dict = {"ok": True, "path": path}
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        vba = [n for n in names if n.lower().endswith("vbaproject.bin")]
        out["vba"] = bool(vba)
        if vba:
            out["vbaSha256"] = sha(z.read(vba[0]))
        pivots = [n for n in names if "pivot" in n.lower() and not n.endswith("/")]
        out["pivotParts"] = {n: sha(z.read(n)) for n in pivots}
        pivot_xml = [
            n
            for n in pivots
            if n.lower().startswith("xl/pivottables/")
            and n.endswith(".xml")
            and "/_rels/" not in n
        ]
        out["pivotXmlFiles"] = pivot_xml
        defs = []
        for n in pivot_xml:
            raw = z.read(n)
            root = ET.fromstring(raw)
            loc = root.find("m:location", NS)
            defs.append(
                {
                    "part": n,
                    "name": root.attrib.get("name"),
                    "location": None
                    if loc is None
                    else {
                        "ref": loc.attrib.get("ref"),
                        "firstHeaderRow": loc.attrib.get("firstHeaderRow"),
                        "firstDataRow": loc.attrib.get("firstDataRow"),
                        "firstDataCol": loc.attrib.get("firstDataCol"),
                        "rowPageCount": loc.attrib.get("rowPageCount"),
                        "colPageCount": loc.attrib.get("colPageCount"),
                    },
                    "cacheId": root.attrib.get("cacheId"),
                    "dataCaption": root.attrib.get("dataCaption"),
                    "attrs": dict(root.attrib),
                }
            )
        out["pivotDefinitions"] = defs
        sheet_rels = [
            n
            for n in names
            if n.lower().startswith("xl/worksheets/_rels/") and n.endswith(".rels")
        ]
        out["worksheetRels"] = {}
        for n in sheet_rels:
            out["worksheetRels"][n] = z.read(n).decode("utf-8", errors="replace")[:4000]

    wb = load_workbook(path, keep_vba=True, data_only=False)
    out["sheets"] = list(wb.sheetnames)
    if "P-Roles" not in wb.sheetnames:
        out["ok"] = False
        out["error"] = "P-Roles missing"
        wb.close()
        return out
    ws = wb["P-Roles"]
    out["pRoles"] = {
        "maxRow": ws.max_row,
        "maxCol": ws.max_column,
        "merged": [str(m) for m in ws.merged_cells.ranges],
        "dimensions": str(ws.dimensions),
        "sheetState": ws.sheet_state,
        "freeze": str(ws.freeze_panes),
        "autoFilter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
        "tables": [t.name for t in (ws.tables or [])],
        "printArea": ws.print_area,
    }
    dvs = []
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            dvs.append(
                {
                    "sqref": str(dv.sqref),
                    "type": dv.type,
                    "formula1": str(dv.formula1) if dv.formula1 is not None else None,
                    "formula2": str(dv.formula2) if dv.formula2 is not None else None,
                }
            )
    out["pRoles"]["dataValidations"] = dvs

    used = []
    for r in range(1, min(ws.max_row or 1, 80) + 1):
        for c in range(1, min(ws.max_column or 1, 20) + 1):
            cell = ws.cell(r, c)
            if cell.value in (None, ""):
                continue
            used.append(
                {
                    "r": r,
                    "c": c,
                    "coord": cell.coordinate,
                    "value": str(cell.value)[:120],
                    "numFmt": cell.number_format,
                    "fontBold": bool(cell.font and cell.font.bold),
                }
            )
    out["pRolesUsedCells"] = used
    out["pRolesUsedCount"] = len(used)

    grid = []
    for r in range(1, 25):
        row = []
        for c in range(1, 12):
            v = ws.cell(r, c).value
            row.append("" if v is None else str(v)[:80])
        grid.append(row)
    out["pRolesGrid24x11"] = grid

    if "Master Sheet" in wb.sheetnames:
        ms = wb["Master Sheet"]
        out["masterHeaders"] = [
            str(ms.cell(1, c).value or "").strip() for c in range(1, 14)
        ]
        last = 1
        for i, row in enumerate(
            ms.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2
        ):
            if row[0] is not None and str(row[0]).strip():
                last = i
        out["masterDataRows"] = last - 1
    wb.close()
    return out


if __name__ == "__main__":
    print(json.dumps(inspect(sys.argv[1]), indent=2))
