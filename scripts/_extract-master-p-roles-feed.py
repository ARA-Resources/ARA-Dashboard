"""Extract P-Roles feed columns from Master Sheet. Read-only. Never saves the workbook."""
import json
import sys
from openpyxl import load_workbook

HEADERS = [
    "Job Requisition ID",
    "Primary Skills",
    "Skill Categorization",
    "Job Management Level",
    "Job Status",
    "Posted",
    "Market Map",
]


def cell(ws, r, c):
    v = ws.cell(r, c).value
    if v is None:
        return ""
    return str(v).strip()


def extract(path: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    if "Master Sheet" not in wb.sheetnames:
        raise RuntimeError("Master Sheet not found")
    ws = wb["Master Sheet"]
    header = [str(x or "").strip() for x in next(ws.iter_rows(min_row=1, max_row=1, max_col=13, values_only=True))]
    wanted = {
        "Job Requisition ID": None,
        "Primary Skills": None,
        "Skill Categorization": None,
        "Job Management Level": None,
        "Job Status": None,
        "Posted": None,
        "Market Map": None,
    }
    for i, name in enumerate(header, start=1):
        if name in wanted and wanted[name] is None:
            wanted[name] = i
    missing = [k for k, v in wanted.items() if v is None]
    if missing:
        raise RuntimeError(f"Missing Master headers: {missing}. Found: {header}")

    rows = []
    for r in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        jr = "" if r[wanted["Job Requisition ID"] - 1] is None else str(r[wanted["Job Requisition ID"] - 1]).strip()
        if not jr:
            continue
        rows.append([
            jr,
            "" if r[wanted["Primary Skills"] - 1] is None else str(r[wanted["Primary Skills"] - 1]).strip(),
            "" if r[wanted["Skill Categorization"] - 1] is None else str(r[wanted["Skill Categorization"] - 1]).strip(),
            "" if r[wanted["Job Management Level"] - 1] is None else str(r[wanted["Job Management Level"] - 1]).strip(),
            "" if r[wanted["Job Status"] - 1] is None else str(r[wanted["Job Status"] - 1]).strip(),
            "" if r[wanted["Posted"] - 1] is None else str(r[wanted["Posted"] - 1]).strip(),
            "" if r[wanted["Market Map"] - 1] is None else str(r[wanted["Market Map"] - 1]).strip(),
        ])
    wb.close()
    return {
        "ok": True,
        "headers": HEADERS,
        "masterHeaders": header,
        "rowCount": len(rows),
        "rows": rows,
    }


def main() -> None:
    if len(sys.argv) < 3:
        print(json.dumps({"ok": False, "error": "usage: extract.py <xlsm> <out.json>"}))
        return
    src, dest = sys.argv[1], sys.argv[2]
    payload = extract(src)
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    print(json.dumps({
        "ok": True,
        "rowCount": payload["rowCount"],
        "headers": payload["headers"],
        "masterHeaders": payload["masterHeaders"],
        "out": dest,
    }))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
