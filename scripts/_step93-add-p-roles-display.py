"""Surgically insert P-Roles Display sheet into an XLSM zip without rewriting pivot/VBA/worksheets."""
from __future__ import annotations

import json
import os
import sys
import zipfile
from xml.sax.saxutils import escape

SHEET_NAME = "P-Roles Display"
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

JMLS = [
    "8-Associate Manager",
    "9-Team Lead/Consultant",
    "10-Senior Analyst",
    "11-Analyst",
    "12-Associate",
]


def countifs_term(status: str, jml_header_cell: str) -> str:
    return (
        f"COUNTIFS('Master Sheet'!$F:$F,sk,'Master Sheet'!$E:$E,ct,"
        f"'Master Sheet'!$G:$G,{jml_header_cell},"
        f"'Master Sheet'!$K:$K,\"{status}\","
        f"'Master Sheet'!$M:$M,IF($B$11=\"All\",\"*\",$B$11),"
        f"'Master Sheet'!$I:$I,IF($B$12=\"All\",\"*\",$B$12))"
    )


def map_for(jml_header_cell: str) -> str:
    def term(status: str, flag: str) -> str:
        return f"IF({flag},{countifs_term(status, jml_header_cell)},0)"

    body = (
        term("Active", "$B$6")
        + "+"
        + term("New", "$B$7")
        + "+"
        + term("Reopen", "$B$8")
        + "+"
        + term("Closed", "$B$9")
    )
    return (
        '_xlfn.IFERROR(_xlfn.MAP(A16:A,B16:B,_xlfn.LAMBDA(sk,ct,IF(OR(sk="",ct=""),"",'
        + body
        + '))),"")'
    )


def inline(text: str) -> str:
    return f'<c t="inlineStr"><is><t>{escape(text)}</t></is></c>'.replace(
        "<c ", "<c r=\"{ref}\" "
    )


def c_inline(ref: str, text: str) -> str:
    return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escape(text)}</t></is></c>'


def c_bool(ref: str, value: bool) -> str:
    return f'<c r="{ref}" t="b"><v>{1 if value else 0}</v></c>'


def c_formula(ref: str, formula: str) -> str:
    return f'<c r="{ref}"><f>{escape(formula)}</f></c>'


def build_sheet_xml() -> bytes:
    unique_pairs = (
        '_xlfn.IFERROR(_xlfn.UNIQUE(_xlfn._xlws.FILTER(_xlfn.HSTACK(\'Master Sheet\'!F2:F,\'Master Sheet\'!E2:E),'
        '(\'Master Sheet\'!B2:B<>"")*'
        '(((\'Master Sheet\'!K2:K="Active")*$B$6)+'
        '((\'Master Sheet\'!K2:K="New")*$B$7)+'
        '((\'Master Sheet\'!K2:K="Reopen")*$B$8)+'
        '((\'Master Sheet\'!K2:K="Closed")*$B$9))*'
        '(IF($B$11="All",((\'Master Sheet\'!M2:M="Yes")+(\'Master Sheet\'!M2:M="-")),'
        '\'Master Sheet\'!M2:M=$B$11))*'
        '(IF($B$12="All",TRUE,\'Master Sheet\'!I2:I=$B$12)))), "")'
    )
    grand_map = (
        '_xlfn.IFERROR(_xlfn.MAP(C16:C,D16:D,E16:E,F16:F,G16:G,'
        '_xlfn.LAMBDA(a,b,c,d,e,IF(AND(a="",b="",c="",d="",e=""),"",N(a)+N(b)+N(c)+N(d)+N(e)))), "")'
    )
    market_list = (
        '_xlfn.IFERROR(_xlfn.VSTACK("All", _xlfn.UNIQUE(_xlfn._xlws.FILTER(\'Master Sheet\'!I2:I,\'Master Sheet\'!I2:I<>""))),"All")'
    )
    headers = ["Primary Skills", "Skill Categorization", *JMLS, "Grand Total"]
    header_cells = "".join(
        c_inline(f"{chr(65+i)}15", h) for i, h in enumerate(headers)
    )
    rows = [
        f'<row r="1">{c_inline("A1", "P-Roles - Google Sheets live display")}{c_inline("Z1", "Market Map list")}</row>',
        f'<row r="2">{c_inline("A2", "Source: Master Sheet live formulas. Excel users: use the P-Roles PivotTable. This tab is for Google Sheets.")}{c_formula("Z2", market_list)}</row>',
        f'<row r="4">{c_inline("A4", "Filters (change these; the table below recalculates)")}</row>',
        f'<row r="5">{c_inline("A5", "Job Status include (TRUE/FALSE). Closed stays available.")}</row>',
        f'<row r="6">{c_inline("A6", "Active")}{c_bool("B6", True)}</row>',
        f'<row r="7">{c_inline("A7", "New")}{c_bool("B7", True)}</row>',
        f'<row r="8">{c_inline("A8", "Reopen")}{c_bool("B8", True)}</row>',
        f'<row r="9">{c_inline("A9", "Closed")}{c_bool("B9", False)}</row>',
        f'<row r="11">{c_inline("A11", "Posted")}{c_inline("B11", "All")}{c_inline("C11", "All | Yes | -")}</row>',
        f'<row r="12">{c_inline("A12", "Market Map")}{c_inline("B12", "All")}{c_inline("C12", "All, or a Market Map value")}</row>',
        f'<row r="14">{c_inline("A14", "Displayed Grand Total (sum of visible row Grand Totals)")}{c_formula("B14", "IFERROR(SUM(H16:H),0)")}</row>',
        f'<row r="15">{header_cells}</row>',
        (
            f'<row r="16">{c_formula("A16", unique_pairs)}'
            f'{c_formula("C16", map_for("C$15"))}'
            f'{c_formula("D16", map_for("D$15"))}'
            f'{c_formula("E16", map_for("E$15"))}'
            f'{c_formula("F16", map_for("F$15"))}'
            f'{c_formula("G16", map_for("G$15"))}'
            f'{c_formula("H16", grand_map)}</row>'
        ),
    ]
    xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{NS_MAIN}" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetPr/>
  <dimension ref="A1:Z16"/>
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="15" topLeftCell="A16" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <sheetFormatPr defaultRowHeight="15"/>
  <cols>
    <col min="1" max="1" width="42" customWidth="1"/>
    <col min="2" max="2" width="28" customWidth="1"/>
    <col min="3" max="8" width="20" customWidth="1"/>
    <col min="26" max="26" width="28" customWidth="1"/>
  </cols>
  <sheetData>
    {''.join(rows)}
  </sheetData>
  <dataValidations count="3">
    <dataValidation type="list" allowBlank="0" showDropDown="0" sqref="B6:B9">
      <formula1>"TRUE,FALSE"</formula1>
    </dataValidation>
    <dataValidation type="list" allowBlank="0" showDropDown="0" sqref="B11">
      <formula1>"All,Yes,-"</formula1>
    </dataValidation>
    <dataValidation type="list" allowBlank="0" showDropDown="0" sqref="B12">
      <formula1>$Z$2:$Z$200</formula1>
    </dataValidation>
  </dataValidations>
  <pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>
</worksheet>
'''
    return xml.encode("utf-8")


def inject(path: str) -> dict:
    sheet_xml = build_sheet_xml()
    tmp = path + ".step93.zip"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        names = zin.namelist()
        if "xl/worksheets/sheet8.xml" in names:
            raise RuntimeError("sheet8.xml already exists")
        workbook = zin.read("xl/workbook.xml").decode("utf-8")
        rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        types = zin.read("[Content_Types].xml").decode("utf-8")
        app = zin.read("docProps/app.xml").decode("utf-8")
        if 'name="P-Roles Display"' in workbook:
            raise RuntimeError("P-Roles Display already present")
        if "rId18" in rels:
            rid = "rId19"
        else:
            rid = "rId18"
        workbook = workbook.replace(
            "</sheets>",
            f'<sheet name="{SHEET_NAME}" sheetId="8" r:id="{rid}"/></sheets>',
            1,
        )
        rels = rels.replace(
            "</Relationships>",
            f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet8.xml"/></Relationships>',
            1,
        )
        types = types.replace(
            "</Types>",
            '<Override PartName="/xl/worksheets/sheet8.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>',
            1,
        )
        if "<vt:i4>7</vt:i4>" in app and 'size="7"' in app:
            app = app.replace("<vt:i4>7</vt:i4>", "<vt:i4>8</vt:i4>", 1)
            app = app.replace(
                '<vt:vector size="7" baseType="lpstr">',
                '<vt:vector size="8" baseType="lpstr">',
                1,
            )
            app = app.replace(
                "</vt:vector></TitlesOfParts>",
                f"<vt:lpstr>{SHEET_NAME}</vt:lpstr></vt:vector></TitlesOfParts>",
                1,
            )
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                data = workbook.encode("utf-8")
            elif info.filename == "xl/_rels/workbook.xml.rels":
                data = rels.encode("utf-8")
            elif info.filename == "[Content_Types].xml":
                data = types.encode("utf-8")
            elif info.filename == "docProps/app.xml":
                data = app.encode("utf-8")
            zout.writestr(info, data)
        zout.writestr("xl/worksheets/sheet8.xml", sheet_xml)
    os.replace(tmp, path)
    return {"ok": True, "sheet": SHEET_NAME, "rid": rid, "part": "xl/worksheets/sheet8.xml"}


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "path required"}))
        return
    try:
        result = inject(sys.argv[1])
        from openpyxl import load_workbook
        wb = load_workbook(sys.argv[1], read_only=True, keep_vba=True)
        result["sheets"] = list(wb.sheetnames)
        wb.close()
        print(json.dumps(result))
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))


if __name__ == "__main__":
    main()
