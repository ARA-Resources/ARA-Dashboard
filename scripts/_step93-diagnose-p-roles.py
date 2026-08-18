"""Diagnose P-Roles layout and Master fields for Step 9.3."""
import json
import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, read_only=False, data_only=False, keep_vba=True)
out = {"ok": True, "sheets": wb.sheetnames}

ms = wb["Master Sheet"]
out["masterHeaders"] = [str(ms.cell(1, c).value or "").strip() for c in range(1, 14)]
last = 1
for i, row in enumerate(ms.iter_rows(min_col=2, max_col=2, min_row=2, values_only=True), start=2):
    if row[0] is not None and str(row[0]).strip():
        last = i
out["masterDataRows"] = last - 1
# sample Posted / Job Status
from collections import Counter
posted = Counter()
status = Counter()
jml = Counter()
for r in range(2, min(last, 50) + 1):
    pass
for r in range(2, last + 1):
    posted[str(ms.cell(r, 13).value or "").strip()] += 1
    status[str(ms.cell(r, 11).value or "").strip()] += 1
    jml[str(ms.cell(r, 7).value or "").strip()] += 1
out["postedCounts"] = dict(posted)
out["statusCounts"] = dict(status)
out["jmlCounts"] = dict(jml)

ps = wb["P-Roles"]
out["pRolesMaxRow"] = ps.max_row
out["pRolesMaxCol"] = ps.max_column
# sample first 12x12 cells
grid = []
for r in range(1, 13):
    row = []
    for c in range(1, 13):
        v = ps.cell(r, c).value
        row.append("" if v is None else str(v)[:80])
    grid.append(row)
out["pRolesTopLeft"] = grid
# last non-empty row/col sample
used = []
for r in range(1, min(ps.max_row or 1, 40) + 1):
    for c in range(1, min(ps.max_column or 1, 20) + 1):
        v = ps.cell(r, c).value
        if v not in (None, ""):
            used.append({"r": r, "c": c, "v": str(v)[:60]})
out["pRolesUsedSample"] = used[:80]
out["pRolesUsedCount"] = len(used)
wb.close()
print(json.dumps(out, indent=2))
