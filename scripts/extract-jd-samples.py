"""
Extract Master Sheet Job Description samples as JSON (read-only).
Usage: python scripts/extract-jd-samples.py <xlsm-path> <out-json>
"""
from __future__ import annotations

import json
import sys
from openpyxl import load_workbook


def cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).replace("\u00a0", " ")


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: python extract-jd-samples.py <workbook> <out.json>", file=sys.stderr)
        return 2

    src = sys.argv[1]
    out = sys.argv[2]

    wb = load_workbook(src, read_only=True, data_only=True, keep_vba=False)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() == "master sheet":
            sheet = wb[name]
            break
    if sheet is None:
        for name in wb.sheetnames:
            if "master" in name.lower():
                sheet = wb[name]
                break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers = [cell_to_text(h).strip() for h in header_row]
    # trim trailing empties
    while headers and not headers[-1]:
        headers.pop()

    job_desc_idx = next(
        (i for i, h in enumerate(headers) if "job" in h.lower() and "description" in h.lower()),
        None,
    )
    if job_desc_idx is None:
        raise SystemExit("Job Description column not found")

    records = []
    for values in rows_iter:
        if values is None:
            continue
        row = {}
        has = False
        for i, h in enumerate(headers):
            if not h:
                continue
            v = values[i] if i < len(values) else None
            text = cell_to_text(v)
            row[h] = text
            if text.strip():
                has = True
        if has:
            records.append(row)

    wb.close()

    payload = {
        "sheetName": sheet.title,
        "headers": [h for h in headers if h],
        "jobDescriptionHeader": headers[job_desc_idx],
        "rowCount": len(records),
        "rows": records,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)

    print(json.dumps({"ok": True, "rows": len(records), "sheet": sheet.title}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
